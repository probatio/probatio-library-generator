from openpyxl import load_workbook
import os

workbook = load_workbook(filename="PROBATIO_I2C_ADDRESS_SPACE.xlsx")

file_h = 'buffer_emulator.txt'
output_file = open(file_h, 'w')

file_generate_array = 'probatio_array.cpp'
output_file_array = open(file_generate_array, 'w')

sheet = workbook.active
count_blocks = 0
count_blocks_size = 0

# output_file.write('/*\nProbatio addresses and definitions:\n*/\n\n')
# output_file_array.write('/*\nProbatio populating array:\n*/\n\n')

blocks_names = str()
buffer_emulator = list()

buffer_emulator.append("2 - Buffer begin")
for row in sheet.iter_rows(values_only=True):
        
    block_address = row[1]
    block_name = row[3]
    block_data_size = row[4]
    is_active = row[5]
    if(block_address != 'ADDRESS' and block_name != None and is_active == 'OK'):
        buffer_emulator.append(str(block_address) + " " + block_name)
        for i in range(int(block_data_size)):
            buffer_emulator.append('#')
        #output_file.write(f'#define BLOCK_{block_name} {block_address}\n')
        #output_file.write(f'#define SIZE_BLOCK_{block_name} {block_data_size}\n\n')
        #count_blocks = count_blocks + 1
        #count_blocks_size = count_blocks_size + int(block_data_size)
        #output_file_array.write(f'Block {block_name.lower()}(BLOCK_{block_name}, SIZE_BLOCK_{block_name}, 0, 0, \"{block_name.lower()}\");\n')
        #blocks_names = ', '.join([blocks_names, f'&{block_name.lower()}'])
        #blocks_names = blocks_names + f', &{block_name.lower()}'
buffer_emulator.append("10 - Buffer end")

buffer_size = count_blocks + count_blocks_size + 2

# output_file.write(f'#define QUANTITY_BLOCKS {count_blocks}\n')
# output_file.write(f'#define BUFFER_SIZE {buffer_size}')

# output_file_array.write(f'Block* blocks[QUANTITY_BLOCKS] = {{{blocks_names}}};')

index = 0
for entry in buffer_emulator:
    index = index + 1
    output_file.write(str(index) + ' => ' + entry + '\n')

output_file.close()
output_file_array.close()